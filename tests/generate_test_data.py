import os

import openpyxl


def create_sheet_simple(wb):
    """Add a sheet named 'simple' with basic cell data and merged cells."""
    ws = wb.active
    ws.title = "simple"

    # Add some simple data
    ws["A1"] = "Header #1"
    ws["B1"] = "Header #2"
    ws["C1"] = "Header #3"

    ws["A2"] = "ABC"
    ws["B2"] = 123.45
    ws["C2"] = "Alice"

    ws["A3"] = "DEF"
    ws["B3"] = 678.0
    ws["C3"] = "Bob"

    ws["C4"] = "Charlie"

    ws["C5"] = "David"

    # Add some merged cells to test merges
    # Merge A4:B5
    ws.merge_cells("A4:B5")
    ws["A4"] = "Merged value"


def create_sheet_comple(wb):
    """Add a sheet named 'complex' with a complex layout (multiple tables, merged cells, different data types)."""
    ws = wb.create_sheet(title="complex")

    # 1. Title block
    ws.merge_cells("A1:E1")
    ws["A1"] = "Financial Report 2026"

    # 2. Table 1: Sales Data (Row 3 to 8)
    ws["A3"] = "Region"
    ws["B3"] = "Q1"
    ws["C3"] = "Q2"
    ws["D3"] = "Q3"
    ws["E3"] = "Q4"

    sales_data = [
        ["North", 12000.50, 15000.75, 13000.00, 18000.25],
        ["South", 9500.00, 11000.50, 10500.20, 12000.80],
        ["East", 14000.10, 16500.30, 15000.00, 21000.50],
        ["West", 11000.00, 12500.00, 13000.00, 14500.00],
    ]
    for row_idx, row_data in enumerate(sales_data, start=4):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Total row
    ws["A8"] = "Total"
    ws["B8"] = "=SUM(B4:B7)"
    ws["C8"] = "=SUM(C4:C7)"
    ws["D8"] = "=SUM(D4:D7)"
    ws["E8"] = "=SUM(E4:E7)"

    # 3. Table 2: Employee Directory (Row 10 to 12)
    ws["A10"] = "Name"
    ws["B10"] = "Role"
    ws["C10"] = "Active"
    ws["D10"] = "Start Date"

    employees = [
        ["Alice", "Engineer", True, "2023-01-15"],
        ["Bob", "Manager", False, "2021-06-01"],
    ]
    for row_idx, row_data in enumerate(employees, start=11):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # 4. Merged metadata block
    ws.merge_cells("B14:D15")
    ws["B14"] = "Confidential - Internal Use Only"


def generate():
    """Generate sample Excel test data for acceptance testing."""
    os.makedirs("tests/data", exist_ok=True)

    wb = openpyxl.Workbook()
    create_sheet_simple(wb)
    create_sheet_comple(wb)

    wb.save("tests/data/sample.xlsx")
    print("Sample Excel created at tests/data/sample.xlsx")


if __name__ == "__main__":
    generate()
