import datetime
import os
import random

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill


def create_sheet_simple(wb):
    """Add a sheet named 'simple' with basic cell data and merged cells."""
    ws = wb.active
    ws.title = "simple"

    # Add some simple data
    ws["A1"] = "Header #1"
    ws["B1"] = "Header #2"
    ws["C1"] = "Header #3"

    ws["A2"] = "ABC"
    ws["B2"] = 123.45
    ws["C2"] = "Alice"

    ws["A3"] = "DEF"
    ws["B3"] = 678.0
    ws["C3"] = "Bob"

    ws["C4"] = "Charlie"

    ws["C5"] = "David"

    # Add some merged cells to test merges
    # Merge A4:B5
    ws.merge_cells("A4:B5")
    ws["A4"] = "Merged value"


def create_sheet_complex(wb):
    """Add a sheet named 'complex' with a complex layout (multiple tables, merged cells, different data types)."""
    ws = wb.create_sheet(title="complex")

    # 1. Title block
    ws.merge_cells("A1:E1")
    ws["A1"] = "Financial Report 2026"

    # 2. Table 1: Sales Data (Row 3 to 8)
    ws["A3"] = "Region"
    ws["B3"] = "Q1"
    ws["C3"] = "Q2"
    ws["D3"] = "Q3"
    ws["E3"] = "Q4"

    sales_data = [
        ["North", 12000.50, 15000.75, 13000.00, 18000.25],
        ["South", 9500.00, 11000.50, 10500.20, 12000.80],
        ["East", 14000.10, 16500.30, 15000.00, 21000.50],
        ["West", 11000.00, 12500.00, 13000.00, 14500.00],
    ]
    for row_idx, row_data in enumerate(sales_data, start=4):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Total row
    ws["A8"] = "Total 2026"
    ws["B8"] = "=SUM(B4:B7)"
    ws["C8"] = "=SUM(C4:C7)"
    ws["D8"] = "=SUM(D4:D7)"
    ws["E8"] = "=SUM(E4:E7)"

    # 3. Table 2: Employee Directory (Row 10 to 12)
    ws["A10"] = "Name"
    ws["B10"] = "Role"
    ws["C10"] = "Active"
    ws["D10"] = "Start Date"

    employees = [
        ["Alice", "Engineer", True, datetime.date(2023, 1, 15)],
        ["Bob", "Manager", False, datetime.date(2021, 6, 1)],
    ]
    for row_idx, row_data in enumerate(employees, start=11):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # 4. Merged metadata block
    ws.merge_cells("B14:D15")
    ws["B14"] = "Confidential - Internal Use Only"


def create_sheet_multi_tables(wb):
    """Add a sheet named 'multi-tables' with styled titles/metadata, and territory sub-tables."""
    ws = wb.create_sheet(title="multi-tables")

    # Add some junk data in the 5 first rows of columns A and B
    dates = ["2026-06-01", "2026-06-02", "2026-06-03", "2026-06-04", "2026-06-05"]
    codes = ["INT-CODE-001", "INT-CODE-002", "INT-CODE-003", "INT-CODE-004", "INT-CODE-005"]
    for row in range(1, 6):
        ws[f"A{row}"] = dates[row - 1]
        ws[f"B{row}"] = codes[row - 1]

    # Hide these first 5 rows
    for row in range(1, 6):
        ws.row_dimensions[row].hidden = True

    # Sheet title in cells B7:F7 merged
    ws.merge_cells("B7:F7")
    ws["B7"] = "Financial Report"

    title_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    title_font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")
    for col in range(2, 7):
        cell = ws.cell(row=7, column=col)
        cell.fill = title_fill
        cell.font = title_font
        cell.alignment = center_align

    # Double row height for row 7 (default is typically 15.0 pt)
    ws.row_dimensions[7].height = 35.0

    # Add some metadata in cells C9:D10
    blue_font = Font(bold=True, color="000080")
    ws["C9"] = "Date"
    ws["C9"].font = blue_font
    ws["D9"] = datetime.date(2026, 6, 23)
    ws["D9"].number_format = "yyyy-mm-dd"

    ws["C10"] = "Fiscal Year"
    ws["C10"].font = blue_font
    ws["D10"] = "2025-2026"

    # Seed the random number generator to ensure static values across runs
    random.seed(42)

    territories = ["North", "South", "East", "West"]
    # We can assign different counts of data rows to each territory:
    data_row_counts = {"North": 3, "South": 4, "East": 2, "West": 3}

    current_row = 12
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    footer_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    footer_font = Font(name="Calibri", size=11, bold=True, color="000000")

    for terr in territories:
        # B{current_row}: Territory name
        ws[f"B{current_row}"] = terr
        ws[f"B{current_row}"].font = Font(name="Calibri", size=12, bold=True, color="1F497D")

        # Headers on current_row + 1 and current_row + 2
        r1 = current_row + 1
        r2 = current_row + 2

        ws.row_dimensions[r1].height = 22.0
        ws.row_dimensions[r2].height = 22.0

        # B{r1} (merged with B{r2}): "Product"
        ws.merge_cells(start_row=r1, start_column=2, end_row=r2, end_column=2)
        ws.cell(row=r1, column=2, value="Product")

        # C{r1} (merged with D{r1}): "2025"
        ws.merge_cells(start_row=r1, start_column=3, end_row=r1, end_column=4)
        ws.cell(row=r1, column=3, value=2025)

        # E{r1} (merged with F{r1}): "2026"
        ws.merge_cells(start_row=r1, start_column=5, end_row=r1, end_column=6)
        ws.cell(row=r1, column=5, value=2026)

        # C{r2} and E{r2}: "Expected"
        ws.cell(row=r2, column=3, value="Expected")
        ws.cell(row=r2, column=5, value="Expected")

        # D{r2} and F{r2}: "Actual"
        ws.cell(row=r2, column=4, value="Actual")
        ws.cell(row=r2, column=6, value="Actual")

        # Style headers
        for r in (r1, r2):
            for col in range(2, 7):
                cell = ws.cell(row=r, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align

        # Data rows
        num_rows = data_row_counts[terr]
        data_start = r2 + 1
        for i in range(num_rows):
            row_idx = data_start + i
            # Product name: "Product A", "Product B", etc.
            ws.cell(row=row_idx, column=2, value=f"Product {chr(65 + i)}")
            # Expected/Actual columns: generate random static numbers
            ws.cell(row=row_idx, column=3, value=random.randint(100, 500))  # nosec B311
            ws.cell(row=row_idx, column=4, value=random.randint(90, 510))  # nosec B311
            ws.cell(row=row_idx, column=5, value=random.randint(110, 550))  # nosec B311
            ws.cell(row=row_idx, column=6, value=random.randint(100, 560))  # nosec B311

        data_end = data_start + num_rows - 1

        # Footer row with =SUM(...) formulae
        footer_row = data_end + 1
        ws.cell(row=footer_row, column=2, value="Total")
        ws.cell(row=footer_row, column=3, value=f"=SUM(C{data_start}:C{data_end})")
        ws.cell(row=footer_row, column=4, value=f"=SUM(D{data_start}:D{data_end})")
        ws.cell(row=footer_row, column=5, value=f"=SUM(E{data_start}:E{data_end})")
        ws.cell(row=footer_row, column=6, value=f"=SUM(F{data_start}:F{data_end})")

        ws.row_dimensions[footer_row].height = 20.0
        for col in range(2, 7):
            cell = ws.cell(row=footer_row, column=col)
            cell.fill = footer_fill
            cell.font = footer_font
            if col > 2:
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

        # Advance to the row after the blank row
        current_row = footer_row + 2


def create_sheet_sub_sections(wb):
    """Add a sheet named 'sub-sections' with a table containing headers, sections (sub-headers, data, sub-footers), and a grand total footer."""
    ws = wb.create_sheet(title="sub-sections")

    # Alignments
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")

    # Fonts
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    sub_header_font = Font(name="Calibri", size=11, bold=True, color="1F497D")
    data_font = Font(name="Calibri", size=11, color="000000")
    sub_footer_font = Font(name="Calibri", size=11, bold=True, color="000000")
    footer_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    # Fills
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    sub_header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    sub_footer_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    footer_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    # Set column widths to look nice
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15

    # 1. Header (Row 2)
    headers = ["Project", "Budget", "Spent", "Remaining"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
    ws.row_dimensions[2].height = 24.0

    # 2. Sections data structure
    sections_data = [
        {
            "name": "Development Department",
            "rows": [
                ["Project Alpha", 50000, 45000],
                ["Project Beta", 75000, 80000],
                ["Project Gamma", 30000, 25000],
            ],
        },
        {
            "name": "Marketing Department",
            "rows": [
                ["Campaign X", 20000, 18000],
                ["Campaign Y", 15000, 16000],
            ],
        },
        {
            "name": "Sales Department",
            "rows": [
                ["Direct Sales", 100000, 95000],
                ["Online Sales", 120000, 110000],
                ["Partner Channels", 60000, 62000],
                ["Retail Outlets", 40000, 38000],
            ],
        },
    ]

    current_row = 3
    sub_footer_rows = []

    for section in sections_data:
        # Sub-header: Merge A{current_row}:D{current_row}
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
        ws.cell(row=current_row, column=1, value=section["name"])
        for col_idx in range(1, 5):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.font = sub_header_font
            cell.fill = sub_header_fill
            cell.alignment = center_align
        ws.row_dimensions[current_row].height = 20.0
        current_row += 1

        data_start = current_row
        for row_data in section["rows"]:
            ws.cell(row=current_row, column=1, value=row_data[0]).alignment = left_align
            ws.cell(row=current_row, column=2, value=row_data[1]).alignment = right_align
            ws.cell(row=current_row, column=3, value=row_data[2]).alignment = right_align
            ws.cell(row=current_row, column=4, value=f"=B{current_row}-C{current_row}").alignment = right_align

            for col_idx in range(1, 5):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.font = data_font
            ws.row_dimensions[current_row].height = 18.0
            current_row += 1
        data_end = current_row - 1

        # Sub-footer
        ws.cell(row=current_row, column=1, value="Total").alignment = left_align
        ws.cell(row=current_row, column=2, value=f"=SUM(B{data_start}:B{data_end})").alignment = right_align
        ws.cell(row=current_row, column=3, value=f"=SUM(C{data_start}:C{data_end})").alignment = right_align
        ws.cell(row=current_row, column=4, value=f"=SUM(D{data_start}:D{data_end})").alignment = right_align

        for col_idx in range(1, 5):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.font = sub_footer_font
            cell.fill = sub_footer_fill
        ws.row_dimensions[current_row].height = 20.0
        sub_footer_rows.append(current_row)
        current_row += 1

    # 3. Grand Total Footer
    ws.cell(row=current_row, column=1, value="Grand Total").alignment = left_align
    # Sum of the sub-footers
    budget_sum_formula = "=" + "+".join(f"B{r}" for r in sub_footer_rows)
    spent_sum_formula = "=" + "+".join(f"C{r}" for r in sub_footer_rows)
    remaining_sum_formula = "=" + "+".join(f"D{r}" for r in sub_footer_rows)

    ws.cell(row=current_row, column=2, value=budget_sum_formula).alignment = right_align
    ws.cell(row=current_row, column=3, value=spent_sum_formula).alignment = right_align
    ws.cell(row=current_row, column=4, value=remaining_sum_formula).alignment = right_align

    for col_idx in range(1, 5):
        cell = ws.cell(row=current_row, column=col_idx)
        cell.font = footer_font
        cell.fill = footer_fill
    ws.row_dimensions[current_row].height = 22.0


def generate():
    """Generate sample Excel test data for acceptance testing."""
    dir_path = os.path.dirname(os.path.abspath(__file__))

    wb = openpyxl.Workbook()
    create_sheet_simple(wb)
    create_sheet_complex(wb)
    create_sheet_multi_tables(wb)
    create_sheet_sub_sections(wb)

    output_path = os.path.join(dir_path, "sample.xlsx")
    wb.save(output_path)
    print(f"Sample Excel created at {output_path}")


if __name__ == "__main__":
    generate()
